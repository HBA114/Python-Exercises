import os                               # Sistemde gerekli kutuphane modulleri bulunamamis ise o modulleri indirebilmek icin sisteme erismemizi saglayan os kutuphanesini ekledim.

try:                                    # pandas kutuphanesini bilgisayarda mevcut ise programa dahil etme bilgisayarda bulunamamis ise indirilmesi icin try-except yapisini kullandim.
    import pandas as pd
except ModuleNotFoundError:
    os.system("pip install pandas")

try:                                    # matplotlib kutuphanesini bilgisayarda mevcut ise programa dahil etme bilgisayarda bulunamamis ise indirilmesi icin try-except yapisini kullandim.
    import matplotlib.pyplot as plt
except ModuleNotFoundError:
    os.system("pip install matplotlib")

try:                                    # seaborn kutuphanesini bilgisayarda mevcut ise programa dahil etme bilgisayarda bulunamamis ise indirilmesi icin try-except yapisini kullandim.
    import seaborn as sns
except ModuleNotFoundError:
    os.system("pip install seaborn")

try:                                    # openpyxl kutuphanesini bilgisayarda mevcut ise programa dahil etme bilgisayarda bulunamamis ise indirilmesi icin try-except yapisini kullandim.
    from openpyxl import Workbook,load_workbook
except ModuleNotFoundError:
    os.system("pip install openpyxl")

try:                                    # math kutuphanesini bilgisayarda mevcut ise programa dahil etme bilgisayarda bulunamamis ise indirilmesi icin try-except yapisini kullandim.
    from math import sqrt
except ModuleNotFoundError:
    os.system("pip install math")

def SatirSayisi(YourWs):                # Excel belgesinde okunacak tablonun satir sayisini bulmak icin SatirSayisi fonksiyonu olusturdum.
    sayac = 0
    for i in YourWs:                    # Excel belgesindeki tablonun her bir satiri icin sayaci artiran bir for dongusu olusturdum.
        sayac = sayac + 1
    return sayac                        # Fonksiyon ciktisi olarak satir sayisini dondurdum.

def SutunSayisi(YourWs):                # Excel belgesinde okunacak tablonun sutun bilgisini bulmak icin SutunSayisi fonksiyonunu olusturdum.
    sayac = 0
    for i in YourWs.columns:            # Excel belgesinde tablonun herbir sutunu icin sayaci artiran bir for dongusu olusturdum.
        sayac = sayac + 1
    return sayac                        # Fonksiyon ciktisi olarak sutun sayisini dondurdum.

def Sutunlar(YourWs):                   # Her bir sutun ayri bir degisken tutacagindan yazdıirirken zorluklarla karsilasmamak icin sutunlari Liste(dizi) haline getiren bir fonksiyon yazdim.
    Sutun_Sayisi = SutunSayisi(YourWs)  # Degiskenleri listeye atayabilmek icin  daha once olusturdugum satir sayisini ve sutun sayisini bulan fonksiyonlari kullanarak satir ve sutun sayisini aldim.
    Satir_Sayisi = SatirSayisi(YourWs)
    Liste_Tum = []                      # Tum degiskenlerin bir dizide tutulabilmesi icin liste yapisi olusturdum.
    for i in range(2,Sutun_Sayisi+1):
        Liste_Degisken = []             # Her bir degiskenin verilerinin tutulabilmesi icin bir liste  yapisi olusturdum.
        for j in range(2,Satir_Sayisi+1):
            Liste_Degisken = Liste_Degisken + [ws.cell(j,i).value]
        Liste_Tum = Liste_Tum + [Liste_Degisken]    # Her bir degiskenin Tum degiskenlerin tutulacagi listeye atanabilmesi icin degiskenin tum verilerini listeye aktardim.
    return Liste_Tum                    # Fonksiyon cikti olarak tum degiskenlerin tutuldugu listeyi dondurdum.

def ListeBoyutu(YourList):              # Listelerin boyutunu bulabilmek icin ListeBoyutu fonksiyonunu olusturdum.
    veriSay = 0                         # Fonksiyon girdisi olarak girilen fonksiyonun eleman sayisini bulabilmek icin bir veri sayaci olusturdum.
    for i in YourList:
        veriSay = veriSay + 1           # for dongusu kullanarak sayaci listenin her elemani icin bir kez artirarak listenin eleman sayisini buldum.
    return veriSay                      # Fonksiyon ciktisi olarak Listenin boyutunu dondurdum.

def Toplam(YourList):                   # Listelerin verilerinin toplamini bulabilmek icin Toplam fonksiyonunu olusturdum.
    toplam = 0                          # Verilerin toplamini bulabilmek icin Toplam degiskeni olusturdum.
    for i in range(ListeBoyutu(YourList)):  # for dongusu kulanarak fonksiyon girdisi olarak allinan listenin elemanlarini toplayan bir dongu olusturdum.
        toplam = toplam + YourList[i]
    return toplam                       # Fonksiyon ciktisi olarak Listenin elemanlarinin toplamini dondurdum.

def OrtAritmetik(YourList):             # Listelerin Aritmetik ortalamasini bulabilmek icin OrtAritmetik fonksiyonunu olusturdum.
    veriSay = ListeBoyutu(YourList)     # Aritmetik ortalamayi hesaplayabilmek icin Listenin eleman sayisini ListeBoyutu fonksiyonu ile bulup bir degiskene atadim.
    if veriSay <= 1:                    # Eğer Listede eleman sayisi 1 ve 1'den az ise Listedeki eleman Listenin aritmetik ortalamasi olacagindan fonksiyon ciktisi olarak Listeyi dondurdum.
        return YourList
    else:                               # Listede eleman sayisi 1'den fazla ise fonksiyon ciktisi olarak Listenin elemanlarinin toplamini veri sayisina bolerek aritmetik ortalamayi dondurdum.
        return Toplam(YourList) / veriSay

def Medyan(YourList):                   # Listelerin Medyan(Ortanca)'sini bulabilmek icin Medyan fonksiyonunu olusturdum.
    YourList.sort()                     # Listedeki verilerin medyanini bulabilmek icin verileri sort() metodu ile kucukten buyuge siraladim.
    listeboyut = ListeBoyutu(YourList)  # Listenin tum elemanlarina ulasabilmek icin Liste boyutunu bulup bir degiskene atadim.
    if listeboyut%2 == 0:               # Listenin eleman sayisi 2'ye tam bolunebiliyor ise yani eleman sayisi cift ise Listenin ortasindaki iki elemanin ortalamasini dondurdum.
        return (YourList[int(listeboyut/2)-1] + YourList[int((listeboyut/2)+1)-1])/2
    else:                               # Listenin eleman sayisi 2'ye tam bolunemiyor ise yani eleman sayisi tek ise Listenin ortasindaki elemani dondurdum.
        return YourList[int(len(YourList)/2)]

def TepeDeger(YourList):                # Listelerin Tepe Deger bulabilmek icin TepeDeger fonksiyonunu olusturdum.
    mod = 0                             # Listenin Tepe degerini bulurken degiskenin kac kez gectigini tutan mod degiskenini olusturdum.
    modval = YourList[0]                # Listenin ilk elemani ile baslayip en cok gecen elemani Tepe deger olarak tutacak modval degiskenini olusturdum.
    for i in range(ListeBoyutu(YourList)):
        counter = 0                     # Listenin her bir elemaninin kac kez gectigini saymak icin counter isimli sayac olusturdum.
        for k in range(ListeBoyutu(YourList)):  # Her bir elemanin kac kez gectigini bulabilmek icin for dongusu icinde tekrar bir dongu olusturdum.
            if YourList[k] == YourList[i] : # Eger ilk dongu ile tutulan eleman ile ikinci dongunun tuttugu eleman birbirine esit ise sayaci artirarak degiskenin kac kez gectigini buldum.
                counter = counter + 1
        if counter > mod :              # Eger kac kez gectigi sayilan eleman onceki elemanin gecme sayisindan daha fazla gecmis ise değer olarak yeni mod ve modval degerini atadim.
            mod = counter
            modval = YourList[i]
    return modval                       # Fonksiyon ciktisi olarak Tepe Degeri tutan modval degiskenini dondurdum.

def DegisimAralıgı(YourList):           # Listelerin Degisim Araligini bulabilmek icin DegisimAraligi fonksiyonunu olusturdum.
    YourList.sort()                     # Listenin en kucuk ve en buyuk elemanlarina kolay ve hizli bir sekilde ulasabilmek icin sort() metodu ile Listeyi siraladim.
    veriSay = ListeBoyutu(YourList)     # Listenin eleman sayisini bularak listenin en son ve ilk elemanlarini yani en buyuk ve en kucuk elemanlarina eriserek
    return YourList[veriSay-1] - YourList[0]    # En buyuk eleman ile en kucuk elemanin farkini bularak Degisim araligini fonksiyon ciktisi olarak dondurdum.

def OrtMutlakSapma(YourList):           # Listelerin Ortalama Mutlak Sapmasini bulabilmek icin OrtMutlakSapma fonksiyonu olusturdum.
    veriSay = ListeBoyutu(YourList)     # Listenin eleman sayisini bulabilmek icin ListeBoyutu fonksiyonunu kullandim.
    toplam = 0                          # Listenin her bir elemaninin aritmetik ortalamadan farkinin toplamini tutacak toplam degiskenini olusturdum.
    AritOrtFark = []                    # Listenin her bir elemani icin elemanin aritmetik ortalamadan farkini tutacak bir liste olusturdum.
    for i in range(veriSay):
        ort = OrtAritmetik(YourList)    # Listenin Aritmetik ortalamasini tutacak bir degisken olusturdum ve OrtAritmetik fonksiyonunu kullanarak bu degiskene listenin aritmetik ortalamasini atadim.
        if ort < YourList[i]:           # Listenin Aritmetik ortalamasi elemandan kucuk ise elemandan aritmetik ortalamayi cikartip farki bu farkin tutulacagi listeye ekledim.
            AritOrtFark = AritOrtFark + [YourList[i] - ort]
        else:                           # Listenin Aritmetik ortalamasi elemandan buyuk ise aritmetik ortalamadan elemani cikartip farki bu farkin tutulacagi listeye ekledim.
            AritOrtFark = AritOrtFark + [ort - YourList[i]]
    for i in range(veriSay):            # Her bir elemanin aritmetik ortalamadan farkini toplayarak toplam degiskenine atadim.
        toplam = toplam + AritOrtFark[i]

    return round((toplam/ veriSay),2)   # Fonksiyon ciktisi olarak toplam degiskeninde tutulan elemanlarin aritmetik ortalamadan farklarini veri sayisina bolerek Ortalama mutlak sapmayi dondurdum. Virgulden sonra kac eleman gorunecegini round fonksiyonu ile belirledim.

def Varyans(YourList):                  # Listelerin Varyansini bulabilmek icin Varyans fonksiyonunu olusturdum.
    ort = OrtAritmetik(YourList)        # Varyansi hesaplarken Aritmetik ortalamadan fayadalanamamiz gerektigi icin aritmetik ortalamayi bularak bir degiskene atadim.                     
    veriSay = ListeBoyutu(YourList)     # Dongude her elemana ulasabilmek ve her biri icin hesaplama yapabilmek icin Listenin boyutunu bularak bir degiskene atadim.
    FarkKare = []                       # Listenin elemanlarinin Aritmetik Ortalamadan Farklarini bularak karelerini tutabilmek icin liste olusturdum.
    for i in range(veriSay):            # Listenin her bir elemani icin Aritmetik ortalamadan farkini bulabilmek icin bir for dongusu olusturdum.
        if ort < YourList[i]:           # Listenin Aritmetik ortalamasi elemandan kucuk ise elemandan aritmetik ortalamayi cikartip farkin karesini alarak farklarin karelerinin tutulacagi listeye ekledim.
            FarkKare = FarkKare + [(YourList[i] - ort)*(YourList[i]-ort)]
        else:                           # Listenin Aritmetik ortalamasi elemandan buyuk ise elemandi aritmetik ortalamadan cikartip farkin karesini alarak farklarin karelerinin tutulacagi listeye ekledim.
            FarkKare = FarkKare + [(ort - YourList[i])*(ort - YourList[i])]
    return OrtAritmetik(FarkKare)       # Farklarin karelerinin tutuldugu listedeki elemanlarin aritmetik ortalamasini alarak listenin varyansini fonksiyon ciktisi olarak dondurdum.

def StandartSapma(YourList):            # Listenin verilerinin standart sapmasini bulabilmek icin StandartSapma fonksiyonunu olusturdum.
    ort = OrtAritmetik(YourList)        # Standart sapmayi hesaplarken Aritmetik ortalamadan fayadalanamamiz gerektigi icin aritmetik ortalamayi bularak bir degiskene atadim.
    toplam = 0                          # Listenin elemanlarinin Aritmetik Ortalamadan Farklarini bularak karelerini toplayabilmek icin toplam degiskenini olusturdum.
    veriSay = ListeBoyutu(YourList)     # Dongude her elemana ulasabilmek ve her biri icin hesaplama yapabilmek icin Listenin boyutunu bularak bir degiskene atadim.
    OrtFarkKare = []                    # Listenin elemanlarinin Aritmetik Ortalamadan Farklarini bularak karelerini tutabilmek icin liste olusturdum.
    for i in range(veriSay):            # Listenin her bir elemani icin Aritmetik ortalamadan farkinin karesini bulabilmek icin bir for dongusu olusturdum.
        if ort < YourList[i]:           # Listenin Aritmetik ortalamasi elemandan kucuk ise elemandan aritmetik ortalamayi cikartip farkin karesini alarak farklarin karelerinin tutulacagi listeye ekledim.
            OrtFarkKare = OrtFarkKare + [(YourList[i] - ort)*(YourList[i] - ort)]
        else:                           # Listenin Aritmetik ortalamasi elemandan buyuk ise elemani aritmetik ortalamadan cikartip farkin karesini alarak farklarin karelerinin tutulacagi listeye ekledim.
            OrtFarkKare = OrtFarkKare + [(ort - YourList[i])*(ort - YourList[i])]
        toplam = toplam + OrtFarkKare[i]    # Listenin elemanlarinin Aritmetik Ortalamadan Farklarinin toplamini bulabilmek icin bulunan farklarin karelerini toplam degiskenine ekledim.
                                        # Fonksiyon ciktisi olarak Elde edilen toplamin eleman sayisinin bir eksigine bolunerek cikan sonucun karekokunu dondurdum.
    return round(sqrt(toplam / (veriSay - 1)),2)    # Karekok bulmak icin math kutuphanesinin sqrt metodunu kullandim. Virgulden sonra gosterilecek basamak sayisini belirlemek icin round fonksiyonunu kullandim.

def DegisimKatsayisi(YourList):         # Listenin verilerinin degisim katsayisini bulabilmek icin DegisimKatsayisi fonksiyonunu olusturdum.
    ss = StandartSapma(YourList)        # Listenin verilerinin degisim katsayisini bulabilmek icin Standart sapmayi hesaplayip bir degiskene atadim.
    ort = OrtAritmetik(YourList)        # Listenin verilerinin degisim katsayisini bulabilmek icin aritmetik ortalamayi hesaplayip bir degiskene atadim.

    return round(((ss / ort)*100),2)    # Degisim katsayisini bulabilmek icin Standart sapmayi aritmetik ortalamaya bolup 100 ile carpip sonucu fonksiyon ciktisi olarak dondurdum. Virgulden sonra gosterilecek basamak sayisini belirlemek icin round fonksiyonunu kullandim.

def CeyreklerAcikligi(YourList):        # Listenin verilerinin Ceyrekler acikligini bulabilmek icin CeyreklerAcikligi fonksiyonunu olusturdum.
    YourList.sort()                     # Listenin verilerinin kucuk olan yarisini ve buyuk olan yarisini ayirabilmek icin sort() metodu ile listeyi kucukten buyuge siraladim.
    veriSay = ListeBoyutu(YourList)     # Dongude her elemana ulasabilmek ve her biri icin hesaplama yapabilmek icin Listenin boyutunu bularak bir degiskene atadim.
    ortaDeger = Medyan(YourList)        # Listenin Medayanini bularak medyandan kucuk elemanlar ve medyandan buyuk elemanlar icin iki adet liste olusturdum.
    AltList = []
    UstList = []
    for i in range(veriSay):            # Listenin her bir elemani icin medyandan kucuk olan elemanlari AltList listesine aktardim.
        if YourList[i] < ortaDeger:
            AltList = AltList + [YourList[i]]
        else:                           # Listenin her bir elemani icin medyandan buyuk olan elemanlari UstList listesine aktardim.
            UstList = UstList + [YourList[i]]
    AltCeyrek = Medyan(AltList)         # AltList listesinin medyanini bularak Alt Ceyregi buldum.
    UstCeyrek = Medyan(UstList)         # UstList listesinin medyanini bularak Ust Ceyregi buldum.

    #print("AltCeyrek = ",AltCeyrek)
    #print("UstCeyrek = ",UstCeyrek)
    return UstCeyrek - AltCeyrek        # Ust Ceyrek ile Alt Ceyregin farkini(Ceyrekler Acikligi) fonksiyon ciktisi olarak dondurdum.

def findPath(fileName):
    print(__file__.replace('data_set.py',''))
    filePath = __file__.replace('data_set.py',fileName)
    return filePath



#########################################################################################

dosyaismi = "gecegunduzsicaklikfarklari.xlsx"
filePath = findPath(dosyaismi)


df = pd.read_excel(filePath)   # Pandas kutuphanesi ile excel belgesi okumak icin read_excel metodunu kullandim.

wb = load_workbook(filePath)   # openpyxl kutuphanesi ile excelden aktif calisma sayfasindaki tablolari hucre hucre ayirarak okuyabilmek icin load_workbook metodunu kullandim
ws = wb.active                          # load_workbook modulu ile yukledigim calisma sayfasini kullanabilmek icin aktiflestirdim.


Satir_Sayisi = SatirSayisi(ws)          # Daha once olsusturdugum SatirSayisi fonksiyonu ile aktif calisma sayfasinin satir sayisini buldum.
Sutun_Sayisi = SutunSayisi(ws)          # Daha once olsusturdugum SutunSayisi fonksiyonu ile aktif calisma sayfasinin sutun sayisini buldum.

Liste = Sutunlar(ws)                    # Daha once verileri listeleyebilmek icin olusturdugum Sutunlar fonksiyonu ile aktif calisma sayfasindaki verileri Liste olarak adlandırdıgım Listeye aktardım

#########################################################################################

for i in range(Sutun_Sayisi-1):         # Her bir degisken(Sutun sayisinin 1 eksigi kadar degiskenimiz var) icin istatistik bilgilerini yardirmamiz gerektiginden for dongusu kullandim.
    print("2021 yili "+str(ws.cell(1,i+2).value) + " gece gunduz sicaklik farki icin veriler...")    # istatistigi yazilan degiskenin ismine tablonun belirli hucresine ulasarak eristim ve kullanicinin hangi degisken icin veri gordugunu ekrana yazdirdim.
    print(str(ws.cell(1,i+2).value) + " gece gunduz sicaklik farki Verileri :\n",Liste[i],end="\n") # Listenin tamamindaki verileri kullanicinin gorebilmesi icin ekrana yazdirdim.
    print(str(ws.cell(1,i+2).value) + " gece gunduz sicaklik farki Aritmetik Ortalamasi : ",OrtAritmetik(Liste[i])) # Aritmetik ortalama'yi ekrana yazdirdim.
    print(str(ws.cell(1,i+2).value) + " gece gunduz sicaklik farki Ortancasi(Medyani) : ",Medyan(Liste[i])) # Medyan'i ekrana yazdirdim.
    print(str(ws.cell(1,i+2).value) + " gece gunduz sicaklik farki Tepe Deger(Mod) : ",TepeDeger(Liste[i])) # Tepe Deger'i ekrana yazdirdim.
    print(str(ws.cell(1,i+2).value) + " gece gunduz sicaklik farki Degisim Araligi : ",DegisimAralıgı(Liste[i]))    # Degisim Araligi'ni ekrana yazdirdim.
    print(str(ws.cell(1,i+2).value) + " gece gunduz sicaklik farki Ortalama Mutlak Sapmasi : ",OrtMutlakSapma(Liste[i]))    # Ortalama Mutlak Sapma'yi ekrana yazdirdim.
    print(str(ws.cell(1,i+2).value) + " gece gunduz sicaklik farki Varyansi : {:.2f}".format(Varyans(Liste[i])))    # Varyans'i ekrana yazdirdim.
    print(str(ws.cell(1,i+2).value) + " gece gunduz sicaklik farki Standart  Sapmasi : ",StandartSapma(Liste[i]))   # Standart Sapma'yi ekrana yazdirdim.
    print(str(ws.cell(1,i+2).value) + " gece gunduz sicaklik farki Degisim Katsayisi : ",DegisimKatsayisi(Liste[i]))    # Degisim Katsayisi'ni ekrana yazdirdim.
    print(str(ws.cell(1,i+2).value) + " gece gunduz sicaklik farki Ceyrekler Acikligi : ",CeyreklerAcikligi(Liste[i]),end="\n\n")   # Ceyrekler Acikligi'ni ekrana yazdirdim.

print("Boxplot Grafigi")
# Boxplot Grafigini cizdirebilmek icin matplotlib kutuphanesinden yararlandim.
# %matplotlib inline

df = pd.read_excel(filePath)   # Excel belgesindeki verileri pandas kutuphanesinin read_excel metodu ile bir degiskene atadim.
plt.title("Gece Gündüz Sıcaklık Farkları Boxplot grafiği")   # Boxplot Grafigine bir baslik tanimladim.


g = sns.boxplot(data=df, width=0.7)     # Grafigi olusturmak icin seaborn kutuphanesinin boxplot metodunu kullanarak Boxplot Grafigi cizdirdim.

plt.show()                              # Boxplot Grafigi penceresinin ekranda gorulebilmesi icin matplotlib kutuphanesinin show() metodunu kullandim.
# 'C:\Python310\python.exe -m pip install --upgrade pip'